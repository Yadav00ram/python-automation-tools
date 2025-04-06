import sys
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ✅ Step 1: Automatically find the main folder (today's date format)
current_folder = os.path.dirname(os.path.abspath(__file__))  # Script's folder
today = datetime.now().strftime("%d-%m-%Y")
main_folder = os.path.join(current_folder, today)

if not os.path.exists(main_folder):
    print(f"No folder found with today's date ({today}).")
    exit()

# ✅ Step 2: Define category folders
categories = ["DONE", "EXIST", "INCORRECT"]
category_paths = {cat: os.path.join(main_folder, cat) for cat in categories}

# ✅ Step 3: Identify all designations from folders
designations = set()
designation_counts = {}

for cat, path in category_paths.items():
    if os.path.exists(path):
        for folder in os.listdir(path):
            if os.path.isdir(os.path.join(path, folder)):
                parts = folder.split()
                if len(parts) > 1:
                    designation = " ".join(parts[1:])  # Remove "Done", "Exist", "Incorrect" prefix
                    designations.add(designation)
                    if designation not in designation_counts:
                        designation_counts[designation] = {"Done": 0, "Exist": 0, "Incorrect": 0}
                    # ✅ Count files inside each designation folder
                    designation_counts[designation][cat.capitalize()] = len(os.listdir(os.path.join(path, folder)))

# ✅ Step 4: Prepare data for Excel report
data = []
total_shared, total_done, total_exist, total_incorrect = 0, 0, 0, 0

for index, designation in enumerate(sorted(designation_counts.keys()), start=1):
    done = designation_counts[designation]["Done"]
    exist = designation_counts[designation]["Exist"]
    incorrect = designation_counts[designation]["Incorrect"]
    total = done + exist + incorrect

    data.append([index, today if index == 1 else "", designation, total, done, exist, incorrect])

    total_shared += total
    total_done += done
    total_exist += exist
    total_incorrect += incorrect

# ✅ Add total row
data.append(["", "", "Total", total_shared, total_done, total_exist, total_incorrect])

# ✅ Step 5: Create DataFrame and save to Excel
df = pd.DataFrame(data, columns=["S. No", "Date", "Designation", "Total Shared", "Resumes Uploaded", "Exist", "Incorrect"])
report_file = os.path.join(current_folder, "Formatted_Report.xlsx")
df.to_excel(report_file, index=False)

# ✅ Step 6: Apply Excel Formatting
wb = load_workbook(report_file)
ws = wb.active

# ✅ Apply Header Styling (Green Background, Bold Text)
header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal="center", vertical="center")

for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = bold_font
    cell.alignment = center_align

# ✅ Apply Borders & Alignment to All Cells
for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border
        cell.alignment = center_align

# ✅ Merge Date Column Cells (2nd Column)
ws.merge_cells(start_row=2, start_column=2, end_row=len(data), end_column=2)
date_cell = ws.cell(row=2, column=2)
date_cell.value = today
date_cell.alignment = center_align

# ✅ Auto Adjust Column Width
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get Column Letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2  # Add extra space

# ✅ Save the formatted Excel file
wb.save(report_file)

print(f"✅ Excel report with formatting & merged date column generated successfully: {report_file}")

sys.exit()
