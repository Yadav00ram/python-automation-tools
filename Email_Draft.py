import os
import pandas as pd
import zipfile
import datetime
import win32com.client  # For Outlook (Windows Only)
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ✅ Step 1: main folder ko dhundh
current_folder = os.path.dirname(os.path.abspath(__file__))
today = datetime.datetime.now().strftime("%d-%m-%Y")
main_folder = os.path.join(current_folder, today)

# ✅ Step 2: main folder h ya nhi
if not os.path.exists(main_folder):
    print(f"folder nhi mila h mujhe aaj ki date ke name se, ya to rename kr aise 31-02-2025 ya fir folder la yha ({today}).")
    exit()

# ✅ Step 3: file path dekho
report_file = os.path.join(current_folder, "Formatted_Report.xlsx")
zip_filename = os.path.join(current_folder, f"{today}.zip")

# ✅ Step 4: pahchaaan kro folders ki
categories = ["DONE", "EXIST", "INCORRECT"]
category_paths = {cat: os.path.join(main_folder, cat) for cat in categories}

# ✅ Step 5: file count krna h beta jii
designation_counts = {}

for cat, path in category_paths.items():
    if os.path.exists(path):
        for folder in os.listdir(path):
            folder_path = os.path.join(path, folder)
            if os.path.isdir(folder_path):
                parts = folder.split()
                if len(parts) > 1:
                    designation = " ".join(parts[1:])
                    if designation not in designation_counts:
                        designation_counts[designation] = {"Done": 0, "Exist": 0, "Incorrect": 0}
                    designation_counts[designation][cat.capitalize()] = len(os.listdir(folder_path))

# ✅ Step 6: folder data ko excel me convert kro
data = []
total_shared, total_done, total_exist, total_incorrect = 0, 0, 0, 0

for index, (designation, counts) in enumerate(sorted(designation_counts.items()), start=1):
    done, exist, incorrect = counts["Done"], counts["Exist"], counts["Incorrect"]
    total = done + exist + incorrect
    data.append([index, today if index == 1 else "", designation, total, done, exist, incorrect])

    total_shared += total
    total_done += done
    total_exist += exist
    total_incorrect += incorrect

# ✅ total row jodo
data.append(["", "", "Total", total_shared, total_done, total_exist, total_incorrect])

# ✅ Step 7: excel sheet banegi ab
df = pd.DataFrame(data, columns=["S. No", "Date", "Designation", "Total Shared", "Resumes Uploaded", "Exist", "Incorrect"])
df.to_excel(report_file, index=False)

# ✅ Step 8: Format Excel File
wb = load_workbook(report_file)
ws = wb.active

# ✅ formate dekhte hai
header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))

for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    for cell in col:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

# ✅ sheet ko sahi krte hai
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# ✅ Merge Date Column
ws.merge_cells(start_row=2, start_column=2, end_row=len(data), end_column=2)

wb.save(report_file)

print(f"👍😎 Excel sheet ban gyi beta👍😎: {report_file}")

# ✅ Step 9: main folder ko zip krege
with zipfile.ZipFile(zip_filename, "w", zipfile.ZIP_DEFLATED) as zipf:
    for root, _, files in os.walk(main_folder):
        for file in files:
            file_path = os.path.join(root, file)
            arcname = os.path.relpath(file_path, main_folder)
            zipf.write(file_path, arcname)

print(f"😒Folder zip ho gya bete😆😍: {zip_filename}")

# ✅ Step 10: data ko html table me bnao 
def generate_html_table(dataframe):
    """Generate a properly formatted HTML table from DataFrame."""
    table_style = """
    <style>
        table {
            width: 100%;
            border-collapse: collapse;
            font-family: Arial, sans-serif;
        }
        th {
            background-color: #92D050;
            font-weight: bold;
            text-align: center;
            border: 1px solid black;
            padding: 5px;
        }
        td {
            border: 1px solid black;
            padding: 5px;
            text-align: center;
        }
    </style>
    """
    
    dataframe.iloc[1:, 1] = "&nbsp;"  # Merge effect for Date column

    return table_style + dataframe.to_html(index=False, border=1, escape=False)

html_table = generate_html_table(df)

# ✅ Step 11:table add kro email me 
to_email = "example@gmail.com"  # Change this
subject = f"Resume Report - {today}"
body = f"""<html>
<body>
<p>Dear Sir/Ma’am,</p>

<p><b>Greetings of the day!</b></p>

<p>I am delighted to inform you that all the resumes have been successfully uploaded as per your instructions. Additionally, I have completed the required details in the attached sheet for your review.</p>

<p><b>Report Summary:</b></p>
{html_table}  <!-- Styled Table Will Be Inserted Here -->

<p>Please find the completed zip file attached for your reference. Should you need any further assistance or clarification, please feel free to reach out.</p>

<p>Looking forward to your valuable feedback.</p>

<p>Best regards,<br>
Your Name</p>
</body>
</html>
"""

# ✅ Step 12: outlook me email draft bnate h
try:
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = to_email
    mail.Subject = subject
    mail.HTMLBody = body  

    # ✅ draft ko open krde
    mail.Display()
    print("✅ isko to band kr de ab😂, outlook open ho gya use dekh🤦‍♂️😉")
except Exception as e:
    print(f"gadbad ho gyi kahi too: {e}")
